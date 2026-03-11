import streamlit as st
import datetime
import io
from modules.theme import page_header, step_badge
from modules import database

BULAN_NAMA = ["Januari","Februari","Maret","April","Mei","Juni",
               "Juli","Agustus","September","Oktober","November","Desember"]

STATUS_OPTIONS = ["⏳ Belum", "✅ Tepat", "⚠️ Terlambat", "⚡ Lebih Cepat", "❌ Tidak Terlaksana"]
STATUS_WAJIB_KETERANGAN = ["⚠️ Terlambat", "⚡ Lebih Cepat", "❌ Tidak Terlaksana"]

def status_color(status: str) -> str:
    if "Tepat"       in status: return "#16a34a"   # hijau gelap — kontras di light
    if "Terlambat"   in status: return "#b45309"   # coklat tembaga
    if "Lebih Cepat" in status: return "#0369a1"   # biru tua — kontras di light
    if "Tidak"       in status: return "#dc2626"   # merah
    return "#57534e"                                # abu gelap untuk Belum

def get_minggu_dari_tanggal(tanggal: datetime.date) -> int:
    d = tanggal.day
    if d <= 7:  return 1
    if d <= 14: return 2
    if d <= 21: return 3
    return 4

def sudah_lewat(tahun: int, bulan: int, minggu: int) -> bool:
    today  = datetime.date.today()
    thn_sk = today.year
    bln_sk = today.month
    mgg_sk = get_minggu_dari_tanggal(today)
    if tahun < thn_sk: return True
    if tahun > thn_sk: return False
    if bulan < bln_sk: return True
    if bulan == bln_sk and minggu < mgg_sk: return True
    return False

def hitung_status_otomatis(bulan_r, minggu_r, bulan_a, minggu_a):
    if bulan_r == bulan_a and minggu_r == minggu_a:
        return "✅ Tepat"
    rencana_global = (bulan_r - 1) * 4 + minggu_r
    aktual_global  = (bulan_a - 1) * 4 + minggu_a
    if aktual_global > rencana_global:
        return "⚠️ Terlambat"
    return "⚡ Lebih Cepat"


def show():
    page_header("📝", "Data Inspeksi", "Evaluasi realisasi vs rencana inspeksi")

    tahun_list = database.get_tahun_tersedia()

    if not tahun_list:
        st.html("""
        <div style="background:#fef3c7;border:1px solid #b45309;border-radius:10px;
                    padding:24px;text-align:center;margin-top:32px;">
            <p style="color:#b45309;font-size:18px;font-family:IBM Plex Mono,monospace;
                      font-weight:700;margin-bottom:8px;">⚠️ Belum Ada Rencana Inspeksi</p>
            <p style="color:#57534e;font-size:13px;font-family:IBM Plex Sans,sans-serif;margin:0;">
                Silakan buat dan simpan rencana inspeksi terlebih dahulu di menu
                <strong style="color:#1c1917;">🗓️ Rencana Inspeksi</strong>.
            </p>
        </div>
        """)
        return

    # ── Step 1: Filter ────────────────────────────────────────────────────────
    step_badge(1, "Pilih Rencana Inspeksi")

    col1, col2, col3 = st.columns(3)
    with col1:
        tahun_pilih = st.selectbox("Tahun", options=tahun_list, key="di_tahun")
    mesin_list = database.get_mesin_by_tahun(tahun_pilih)
    with col2:
        mesin_pilih = st.selectbox("Mesin", options=mesin_list, key="di_mesin")
    versi_list = database.get_versi_rencana(mesin_pilih, tahun_pilih)
    versi_options = [
        f"Versi {len(versi_list)-i} — disimpan {v['created_at']} (n={v['n']}x/bulan)"
        for i, v in enumerate(versi_list)
    ]
    with col3:
        versi_idx = st.selectbox(
            "Versi Rencana",
            options=list(range(len(versi_options))),
            format_func=lambda x: versi_options[x],
            key="di_versi",
        )

    rencana_id  = versi_list[versi_idx]["id"]
    jadwal_list = database.get_jadwal_by_rencana(rencana_id)

    if not jadwal_list:
        st.info("Tidak ada jadwal untuk rencana ini.")
        return

    st.markdown("<hr style='border-color:#e7e0d6;margin:16px 0'>", unsafe_allow_html=True)

    # ── Step 2: Input Realisasi ───────────────────────────────────────────────
    step_badge(2, "Input Realisasi Inspeksi")

    # POIN 6: warna teks instruksi kontras di light mode
    st.markdown(
        "<p style='color:#57534e;font-size:12px;font-family:IBM Plex Mono,monospace;"
        "margin-bottom:4px;'>"
        "Centang <span style='color:#dc2626;font-weight:700;'>✗ Tidak Terlaksana</span> "
        "jika inspeksi tidak dilakukan. "
        "Keterangan <span style='color:#dc2626;font-weight:600;'>wajib diisi</span> "
        "jika status bukan Tepat.</p>",
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)
    h1, h2, h3, h4, h5, h6 = st.columns([2, 1.5, 2, 2, 1, 3])
    for col, lbl in zip(
        [h1, h2, h3, h4, h5, h6],
        ["RENCANA", "TDK TERLAKSANA", "BULAN AKTUAL", "MINGGU AKTUAL", "STATUS", "KETERANGAN"]
    ):
        col.markdown(
            f"<p style='color:#78716c;font-size:10px;font-family:IBM Plex Mono,monospace;"
            f"letter-spacing:1px;margin-bottom:2px;'>{lbl}</p>",
            unsafe_allow_html=True,
        )

    for jdw in jadwal_list:
        jid = jdw["jadwal_id"]
        if f"di_ket_{jid}" not in st.session_state:
            st.session_state[f"di_ket_{jid}"] = jdw["keterangan"] or ""
        if f"di_tdk_{jid}" not in st.session_state:
            st.session_state[f"di_tdk_{jid}"] = (
                jdw["status"] == "❌ Tidak Terlaksana" if jdw["status"] else False
            )

    for jdw in jadwal_list:
        jid    = jdw["jadwal_id"]
        bulan  = jdw["bulan"]
        minggu = jdw["minggu"]
        label_rencana = f"{BULAN_NAMA[bulan-1]} Minggu {minggu}"

        bln_default = (jdw["bulan_aktual"]  or bulan)  - 1
        mgg_default = (jdw["minggu_aktual"] or minggu) - 1

        c1, c2, c3, c4, c5, c6 = st.columns([2, 1.5, 2, 2, 1, 3])

        with c1:
            # POIN 6: label rencana "Januari Minggu 1" pakai warna gelap agar terbaca
            st.markdown(
                f"<p style='color:#1c1917;font-family:IBM Plex Mono,monospace;"
                f"font-size:12px;padding-top:8px;font-weight:600;'>{label_rencana}</p>",
                unsafe_allow_html=True,
            )

        with c2:
            tidak_terlaksana = st.checkbox(
                "Tidak Terlaksana",
                key=f"di_tdk_{jid}",
                label_visibility="collapsed",
            )

        if tidak_terlaksana:
            with c3:
                st.markdown(
                    "<p style='color:#78716c;font-family:IBM Plex Mono,monospace;"
                    "font-size:11px;padding-top:8px;'>—</p>",
                    unsafe_allow_html=True,
                )
            with c4:
                st.markdown(
                    "<p style='color:#78716c;font-family:IBM Plex Mono,monospace;"
                    "font-size:11px;padding-top:8px;'>—</p>",
                    unsafe_allow_html=True,
                )
            with c5:
                st.markdown(
                    "<p style='color:#dc2626;font-family:IBM Plex Mono,monospace;"
                    "font-size:11px;padding-top:8px;font-weight:700;'>❌</p>",
                    unsafe_allow_html=True,
                )
            with c6:
                st.text_input(
                    "ket", placeholder="Wajib diisi — alasan tidak terlaksana...",
                    key=f"di_ket_{jid}", label_visibility="collapsed",
                )

        else:
            with c3:
                bln_ak = st.selectbox(
                    "bln", options=list(range(1, 13)),
                    format_func=lambda x: BULAN_NAMA[x-1],
                    index=bln_default,
                    key=f"di_bln_{jid}",
                    label_visibility="collapsed",
                )

            with c4:
                mgg_ak = st.selectbox(
                    "mgg", options=[1, 2, 3, 4],
                    format_func=lambda x: f"Minggu {x}",
                    index=mgg_default,
                    key=f"di_mgg_{jid}",
                    label_visibility="collapsed",
                )

            status_preview = hitung_status_otomatis(bulan, minggu, bln_ak, mgg_ak)
            warna = status_color(status_preview)

            with c5:
                icon = "✅" if "Tepat" in status_preview else (
                       "⚠️" if "Terlambat" in status_preview else "⚡")
                st.markdown(
                    f"<p style='color:{warna};font-family:IBM Plex Mono,monospace;"
                    f"font-size:11px;padding-top:8px;font-weight:700;'>{icon}</p>",
                    unsafe_allow_html=True,
                )

            with c6:
                ket_wajib   = status_preview in STATUS_WAJIB_KETERANGAN
                placeholder = "Wajib diisi..." if ket_wajib else "Opsional..."
                st.text_input(
                    "ket", placeholder=placeholder,
                    key=f"di_ket_{jid}", label_visibility="collapsed",
                )

    st.markdown("<hr style='border-color:#e7e0d6;margin:24px 0'>", unsafe_allow_html=True)

    # ── Step 3: Simpan ────────────────────────────────────────────────────────
    step_badge(3, "Simpan Realisasi")

    if st.button("💾 Simpan Realisasi", type="primary"):
        error_list     = []
        realisasi_list = []

        for jdw in jadwal_list:
            jid              = jdw["jadwal_id"]
            tidak_terlaksana = st.session_state.get(f"di_tdk_{jid}", False)
            ket              = st.session_state.get(f"di_ket_{jid}", "").strip()

            if tidak_terlaksana:
                status = "❌ Tidak Terlaksana"
                bln_ak = None
                mgg_ak = None
                if not ket:
                    error_list.append(
                        f"{BULAN_NAMA[jdw['bulan']-1]} Minggu {jdw['minggu']} "
                        f"(❌ Tidak Terlaksana) — keterangan wajib diisi"
                    )
                    continue
            else:
                bln_ak = st.session_state.get(f"di_bln_{jid}", jdw["bulan"])
                mgg_ak = st.session_state.get(f"di_mgg_{jid}", jdw["minggu"])
                status = hitung_status_otomatis(jdw["bulan"], jdw["minggu"], bln_ak, mgg_ak)
                if status in STATUS_WAJIB_KETERANGAN and not ket:
                    error_list.append(
                        f"{BULAN_NAMA[jdw['bulan']-1]} Minggu {jdw['minggu']} "
                        f"({status}) — keterangan wajib diisi"
                    )
                    continue

            realisasi_list.append({
                "jadwal_id"    : jid,
                "mesin"        : mesin_pilih,
                "tahun"        : tahun_pilih,
                "bulan_aktual" : bln_ak,
                "minggu_aktual": mgg_ak,
                "status"       : status,
                "keterangan"   : ket,
            })

        if error_list:
            st.error("⚠️ Keterangan wajib diisi untuk:\n" + "\n".join(f"• {e}" for e in error_list))
        else:
            database.simpan_realisasi_batch(realisasi_list)
            st.success("✅ Realisasi berhasil disimpan!")
            st.rerun()

    st.markdown("<hr style='border-color:#e7e0d6;margin:16px 0'>", unsafe_allow_html=True)

    # ── Step 4: Rekap Evaluasi ────────────────────────────────────────────────
    step_badge(4, "Rekap Evaluasi")

    jadwal_list = database.get_jadwal_by_rencana(rencana_id)
    total       = len(jadwal_list)
    n_tepat     = sum(1 for j in jadwal_list if j["status"] and "Tepat"       in j["status"])
    n_terlambat = sum(1 for j in jadwal_list if j["status"] and "Terlambat"   in j["status"])
    n_cepat     = sum(1 for j in jadwal_list if j["status"] and "Lebih Cepat" in j["status"])
    n_tidak     = sum(1 for j in jadwal_list if j["status"] and "Tidak"       in j["status"])
    n_belum     = sum(1 for j in jadwal_list if not j["status"] or "Belum"    in j["status"])
    terlaksana  = n_tepat + n_terlambat + n_cepat
    pct         = lambda x: f"{x/total*100:.1f}%" if total > 0 else "0%"

    c1, c2, c3, c4, c5 = st.columns(5)
    def metric_box(col, label, nilai, warna, pct_val=""):
        col.markdown(f"""
        <div style="background:#fef9f0;border:1px solid #e7e0d6;border-radius:8px;
                    padding:14px;text-align:center;">
            <p style="color:{warna};font-size:22px;font-weight:700;
                      font-family:IBM Plex Mono,monospace;margin:0;">{nilai}</p>
            <p style="color:#78716c;font-size:10px;font-family:IBM Plex Mono,monospace;
                      margin:4px 0 0 0;letter-spacing:1px;">{label}</p>
            <p style="color:{warna};font-size:11px;opacity:0.85;margin:2px 0 0 0;">{pct_val}</p>
        </div>
        """, unsafe_allow_html=True)

    metric_box(c1, "TEPAT",            n_tepat,     "#16a34a", pct(n_tepat))
    metric_box(c2, "TERLAMBAT",        n_terlambat, "#b45309", pct(n_terlambat))
    metric_box(c3, "LEBIH CEPAT",      n_cepat,     "#0369a1", pct(n_cepat))
    metric_box(c4, "TIDAK TERLAKSANA", n_tidak,     "#dc2626", pct(n_tidak))
    metric_box(c5, "BELUM",            n_belum,     "#57534e", pct(n_belum))

    st.markdown("<br>", unsafe_allow_html=True)
    pct_terlaksana = terlaksana / total * 100 if total > 0 else 0
    st.markdown(f"""
    <div style="margin-bottom:20px;">
        <div style="display:flex;justify-content:space-between;margin-bottom:6px;">
            <span style="color:#57534e;font-size:11px;font-family:IBM Plex Mono,monospace;">
                TINGKAT REALISASI</span>
            <span style="color:#1c1917;font-size:11px;font-family:IBM Plex Mono,monospace;
                         font-weight:700;">{terlaksana}/{total} ({pct_terlaksana:.1f}%)</span>
        </div>
        <div style="background:#e7e0d6;border-radius:4px;height:8px;">
            <div style="background:linear-gradient(90deg,#16a34a,#0369a1);border-radius:4px;
                        height:8px;width:{pct_terlaksana:.1f}%;"></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<hr style='border-color:#e7e0d6;margin:8px 0 20px 0'>", unsafe_allow_html=True)

    # ── Step 5: Tabel Detail ──────────────────────────────────────────────────
    step_badge(5, "Tabel Detail Rencana vs Realisasi")

    status_filter = st.selectbox(
        "Filter Status", options=["Semua Status"] + STATUS_OPTIONS,
        key="di_filter_status",
    )
    data_tampil = jadwal_list
    if status_filter != "Semua Status":
        data_tampil = [j for j in data_tampil
                       if (j["status"] or "⏳ Belum") == status_filter]

    header_css = ("color:#b45309;font-size:10px;font-family:IBM Plex Mono,monospace;"
                  "letter-spacing:1px;padding:8px 10px;border-bottom:1px solid #d6cfc4;"
                  "text-align:center;background:#fef3c7;")
    cell_css   = ("color:#1c1917;font-size:12px;font-family:IBM Plex Mono,monospace;"
                  "padding:7px 10px;text-align:center;border-bottom:1px solid #ede8e0;")

    rows_html = ""
    for j in data_tampil:
        rencana_lbl = f"{BULAN_NAMA[j['bulan']-1]} Minggu {j['minggu']}"
        aktual_lbl  = (f"{BULAN_NAMA[j['bulan_aktual']-1]} Minggu {j['minggu_aktual']}"
                       if j["bulan_aktual"] and j["minggu_aktual"] else "-")
        status = j["status"] or "⏳ Belum"
        ket    = j["keterangan"] or "-"
        warna  = status_color(status)
        rows_html += f"""
        <tr>
          <td style="{cell_css};text-align:left;border-right:1px solid #d6cfc4;font-weight:600;">{rencana_lbl}</td>
          <td style="{cell_css}">{aktual_lbl}</td>
          <td style="{cell_css};color:{warna};font-weight:700;">{status}</td>
          <td style="{cell_css};text-align:left;color:#57534e;">{ket}</td>
        </tr>
        """

    if not rows_html:
        rows_html = f'<tr><td colspan="4" style="{cell_css};color:#78716c;">Tidak ada data.</td></tr>'

    st.html(f"""
    <div style="overflow-x:auto;margin-top:12px;">
      <table style="border-collapse:collapse;background:#faf8f5;width:100%;
                    font-family:IBM Plex Mono,monospace;">
        <thead>
          <tr>
            <th style="{header_css};text-align:left;border-right:1px solid #d6cfc4;">Rencana</th>
            <th style="{header_css};">Realisasi</th>
            <th style="{header_css};">Status</th>
            <th style="{header_css};text-align:left;">Keterangan</th>
          </tr>
        </thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    <p style="color:#78716c;font-size:10px;font-family:IBM Plex Mono,monospace;margin-top:8px;">
      Menampilkan {len(data_tampil)} dari {len(jadwal_list)} jadwal
    </p>
    """)

    st.markdown("<hr style='border-color:#e7e0d6;margin:20px 0'>", unsafe_allow_html=True)

    # ── Step 6: Export Excel ──────────────────────────────────────────────────
    step_badge(6, "Export Rekap ke Excel")

    if st.button("📥 Export Excel", type="primary"):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb  = openpyxl.Workbook()
            ws  = wb.active
            ws.title = f"Rekap {mesin_pilih} {tahun_pilih}"

            thin   = Side(style="thin", color="d6cfc4")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            # Judul
            ws.merge_cells("A1:F1")
            ws["A1"] = f"REKAP INSPEKSI — {mesin_pilih.upper()} — TAHUN {tahun_pilih}"
            ws["A1"].font      = Font(name="Consolas", bold=True, size=13, color="b45309")
            ws["A1"].fill      = PatternFill("solid", fgColor="fef3c7")
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            ws.merge_cells("A2:F2")
            ws["A2"] = f"Versi: {versi_options[versi_idx]}"
            ws["A2"].font      = Font(name="Consolas", size=10, color="78716c")
            ws["A2"].fill      = PatternFill("solid", fgColor="fef9f0")
            ws["A2"].alignment = Alignment(horizontal="center")
            ws.append([])

            # Header tabel
            hdr_row = ws.max_row + 1
            for ci, hdr in enumerate(
                ["No", "Rencana", "Realisasi", "Status", "Keterangan", "Diupdate"], 1
            ):
                cell = ws.cell(row=hdr_row, column=ci, value=hdr)
                cell.font      = Font(name="Consolas", bold=True, size=10, color="b45309")
                cell.fill      = PatternFill("solid", fgColor="fef3c7")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border
            ws.row_dimensions[hdr_row].height = 20

            status_colors_map = {
                "✅ Tepat"            : "16a34a",
                "⚠️ Terlambat"        : "b45309",
                "⚡ Lebih Cepat"      : "0369a1",
                "❌ Tidak Terlaksana" : "dc2626",
                "⏳ Belum"            : "57534e",
            }

            for no, j in enumerate(jadwal_list, 1):
                rencana_lbl = f"{BULAN_NAMA[j['bulan']-1]} Minggu {j['minggu']}"
                aktual_lbl  = (f"{BULAN_NAMA[j['bulan_aktual']-1]} Minggu {j['minggu_aktual']}"
                               if j["bulan_aktual"] and j["minggu_aktual"] else "-")
                status  = j["status"] or "⏳ Belum"
                s_color = status_colors_map.get(status, "57534e")
                row_data = [no, rencana_lbl, aktual_lbl, status,
                            j["keterangan"] or "-", j["updated_at"] or "-"]
                ws.append(row_data)
                dr = ws.max_row
                ws.row_dimensions[dr].height = 18
                for ci, _ in enumerate(row_data, 1):
                    cell = ws.cell(row=dr, column=ci)
                    cell.fill      = PatternFill("solid", fgColor="faf8f5")
                    cell.alignment = Alignment(
                        horizontal="left" if ci in [2, 3, 5] else "center",
                        vertical="center",
                    )
                    cell.border = border
                    cell.font   = Font(
                        name="Consolas", size=10,
                        color=s_color if ci == 4 else ("57534e" if ci == 5 else "1c1917"),
                        bold=(ci == 4),
                    )

            # Ringkasan
            ws.append([])
            for label, nilai, warna in [
                ("Total Jadwal",      total,                  "1c1917"),
                ("Tepat",             n_tepat,                "16a34a"),
                ("Terlambat",         n_terlambat,            "b45309"),
                ("Lebih Cepat",       n_cepat,                "0369a1"),
                ("Tidak Terlaksana",  n_tidak,                "dc2626"),
                ("Belum",             n_belum,                "57534e"),
                ("Tingkat Realisasi", f"{pct_terlaksana:.1f}%","16a34a"),
            ]:
                r = ws.max_row + 1
                ws.cell(row=r, column=1, value=label).font = \
                    Font(name="Consolas", size=10, color="78716c")
                ws.cell(row=r, column=2, value=str(nilai)).font = \
                    Font(name="Consolas", bold=True, size=10, color=warna)

            for i, w in enumerate([6, 22, 22, 22, 40, 20], 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = f"rekap_inspeksi_{mesin_pilih.replace(' ','_')}_{tahun_pilih}.xlsx"
            st.download_button(
                label="⬇️ Download Excel", data=buf, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ImportError:
            st.error("❌ Tambahkan 'openpyxl' ke requirements.txt")
        except Exception as e:
            st.error(f"❌ Gagal export: {e}")